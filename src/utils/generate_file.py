#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/03/15
# @Author  : MementoMori
# @File    : generate_file.py
# @Software: PyCharm

import os
import random
from pprint import pprint
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
import matplotlib.pyplot as plt
from src.utils.constant import QuestionType, QuestionLevel
import xml.etree.ElementTree as ET


def generate_xml_from_query(results, table_name):
    """

    Args:
        results:
        table_name:

    Returns:

    """
    # 假设results是从数据库查询得到的一个二维列表，每行代表一条记录，每列对应一个字段
    root = ET.Element("root")

    for row in results:
        record = ET.SubElement(root, "record")

        for i, field_value in enumerate(row):
            # 假设你已经有了字段名的列表（这里简化假设它们与顺序一致）
            field_name = "field{}".format(i + 1)  # 替换为实际的字段名
            field_element = ET.SubElement(record, field_name)
            field_element.text = str(field_value)

    # 将XML树转换为字符串
    xml_data = ET.tostring(root, encoding="utf-8", method="xml").decode("utf-8")
    # 写入到文件（可选）
    with open(f"{table_name}.xml", "w") as xml_file:
        xml_file.write(xml_data)

    return xml_data


def generate_excel(table_list, output_path=None, file_name=None):
    """
    生成excel
    Args:
        table_list:
        output_path:
        file_name:

    Returns:

    """
    df = pd.DataFrame(table_list)
    wb = Workbook()
    wb.active = wb.create_sheet(title="Sheet1")
    wb.active.append(df.columns.tolist())
    for row in df.itertuples(index=False):
        wb.active.append(row)
    output_filename = os.path.join(output_path, file_name)
    # f"{output_filename}.xlsx"
    wb.save(output_filename)


def max_column_width(data, column_index):
    """
    获取某列中最长字符串的长度
    Args:
        data:
        column_index:

    Returns:

    """
    return max([len(str(row[column_index])) for row in data])


def generate_pdf(table_list, output_path=None, file_name=None):
    """
    生成pdf
    Args:
        table_list:
        output_path:
        file_name:

    Returns:

    """
    # pdf = FPDF()
    # PAGE_WIDTH = 190  # 单位毫米，此处仅为示例，实际值应为扣除页边距后的宽度
    pdf = FPDF("L")
    PAGE_WIDTH = 270  # 单位毫米，此处仅为示例，实际值应为扣除页边距后的宽度

    simfang_path = "src/static/simfang.ttf"
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)  # 自动分页
    pdf.add_font('simfang', '', simfang_path, uni=True)
    pdf.set_font('simfang', '', 14)

    column_titles = list(table_list[0].keys())
    user_data = [list(row.values()) for row in table_list]

    # 计算每列的最大宽度
    max_widths = [max_column_width(user_data, i) for i in range(len(column_titles))]

    column_widths = [int(PAGE_WIDTH * w / sum(max_widths)) for w in max_widths]
    # 将数据写入PDF
    for i, (title, width) in enumerate(zip(column_titles, column_widths)):
        pdf.cell(width, 10, txt=title, border=1, align='C' if i == 0 else 'L', ln=(i == len(table_list[0]) - 1))
    for record in user_data:
        for col_idx, (value, width) in enumerate(zip(record, column_widths)):
            pdf.cell(width, 10, txt=str(value), border=1, ln=(col_idx == len(record) - 1),
                     align='C' if col_idx == 0 else 'L')

    # 指定要保存的路径
    output_path = "E:\\LOCAL_DOWNLOAD_FILE_PATH\\file.pdf"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    pdf.output(output_path)


class Plotter:
    def __init__(self, figure_size=(10, 6)):
        self.figure_size = figure_size

    @staticmethod
    def configure_plot(title, x_label=None, y_label=None):
        plt.title(title)
        if x_label:
            plt.xlabel(x_label)
        if y_label:
            plt.ylabel(y_label)
        plt.tight_layout()

    def create_line_plot(self, x_data, y_data, x_label='X-axis', y_label='Y-axis', title='Line Plot'):
        plt.figure(figsize=self.figure_size)
        plt.plot(x_data, y_data)
        self.configure_plot(title, x_label, y_label)
        plt.show()

    def create_scatter_plot(self, x_data, y_data, x_label='X-axis', y_label='Y-axis', title='Scatter Plot'):
        plt.figure(figsize=self.figure_size)
        plt.scatter(x_data, y_data)
        self.configure_plot(title, x_label, y_label)
        plt.show()

    def create_bar_chart(self, categories, values, x_label='Categories', y_label='Values', title='Bar Chart'):
        plt.figure(figsize=self.figure_size)
        plt.bar(categories, values)
        plt.xticks(rotation=45)
        self.configure_plot(title, x_label, y_label)
        plt.show()

    def create_pie_chart(self, sizes, labels, title='Pie Chart', autopct='%1.1f%%'):
        plt.figure(figsize=self.figure_size)
        plt.pie(sizes, labels=labels, autopct=autopct, startangle=90)
        plt.axis('equal')
        self.configure_plot(title)
        plt.show()

    def create_histogram(self, data, bins=10, x_label='Value Range', y_label='Frequency', title='Histogram'):
        plt.figure(figsize=self.figure_size)
        plt.hist(data, bins=bins, edgecolor='black')
        self.configure_plot(title, x_label, y_label)
        plt.show()


from src.utils.sql_config import sql_handle


class Question:
    def __init__(self, question_id, question_text, answer, question_type, options=None):
        self.question_id = question_id
        self.question_text = question_text
        self.options = options
        self.answer = answer
        self.question_type = question_type


class ExamPaperGenerator:
    def __init__(self, table_name):
        self.table_name = table_name

    async def load_single_choice_bank(self):
        single_choices = await self.load_question_bank(QuestionType.SingleChoice)
        question_bank = []
        for row in single_choices:
            question_id = row.get('id')
            question_text = row.get('question')
            options = row.get('options')
            answer = row.get('answer')
            question_type = row.get('type')
            question = Question(question_id, question_text, answer, question_type, options)
            question_bank.append(question)
        return question_bank

    async def load_multiple_choice_bank(self):
        multiple_choices = await self.load_question_bank(QuestionType.MultipleChoice)
        question_bank = []
        for row in multiple_choices:
            question_id = row.get('id')
            question_text = row.get('question')
            options = row.get('options')
            answer = row.get('answer')
            question_type = row.get('type')
            question = Question(question_id, question_text, answer, question_type, options)
            question_bank.append(question)
        return question_bank

    async def load_fill_bank(self):
        fills = await self.load_question_bank(QuestionType.Fill)
        question_bank = []
        for row in fills:
            question_id = row.get('id')
            question_text = row.get('question')
            answer = row.get('answer')
            question_type = row.get('type')
            question = Question(question_id, question_text, answer, question_type)
            question_bank.append(question)
        return question_bank

    async def load_judge_bank(self):
        judges = await self.load_question_bank(QuestionType.Judge)
        question_bank = []
        for row in judges:
            question_id = row.get('id')
            question_text = row.get('question')
            options = row.get('options')
            answer = row.get('answer')
            question_type = row.get('type')
            question = Question(question_id, question_text, answer, question_type, options)
            question_bank.append(question)
        return question_bank

    async def load_short_answer_bank(self):
        short_answers = await self.load_question_bank(QuestionType.ShortAnswer)
        question_bank = []
        for row in short_answers:
            question_id = row.get('id')
            question_text = row.get('question')
            answer = row.get('answer')
            question_type = row.get('type')
            question = Question(question_id, question_text, answer, question_type)
            question_bank.append(question)
        return question_bank

    async def load_question_bank(self, question_type, limit=100, is_desc=True):
        # desc 降序
        question_bank = await sql_handle.select(self.table_name, conditions={"type": question_type},
                                                order_by={"update_time": is_desc}, limit=limit)
        return question_bank

    async def generate_exam(self, question_type_counts):
        exam_questions = []
        for question_type, question_num in question_type_counts.items():
            if question_type == "single_choice":
                single_choice_bank = await self.load_single_choice_bank()
                random.shuffle(single_choice_bank)
                exam_questions.extend(single_choice_bank[:question_num])
            if question_type == "multiple_choice":
                multiple_choice_bank = await self.load_multiple_choice_bank()
                random.shuffle(multiple_choice_bank)
                exam_questions.extend(multiple_choice_bank[:question_num])
            if question_type == "fill":
                fill_bank = await self.load_fill_bank()
                random.shuffle(fill_bank)
                exam_questions.extend(fill_bank[:question_num])
            if question_type == "judge":
                judge_bank = await self.load_judge_bank()
                random.shuffle(judge_bank)
                exam_questions.extend(judge_bank[:question_num])
            if question_type == "short_answer":
                short_answer_bank = await self.load_short_answer_bank()
                random.shuffle(short_answer_bank)
                exam_questions.extend(short_answer_bank[:question_num])

        return exam_questions

    def print_exam(self, exam):
        for i, question in enumerate(exam, start=1):
            print(f"Question {i}: {question.question_text}")
            for j, option in enumerate(question.options, start=1):
                print(f"{chr(65 + j - 1)}. {option}")
            print("\n")


if __name__ == '__main__':
    # plotter = Plotter()
    # # 示例数据
    # x = range(10)
    # y = [i ** 2 for i in x]
    # categories = ['A', 'B', 'C', 'D', 'E']
    # values = [30, 25, 40, 35, 20]
    # sizes = [30, 25, 40, 35]
    # labels = ['Category A', 'Category B', 'Category C', 'Category D']

    # 使用示例
    # plotter.create_line_plot(x, y)
    # plotter.create_scatter_plot(x, y)
    # plotter.create_bar_chart(categories, values)
    # plotter.create_pie_chart(sizes, labels)
    # plotter.create_histogram(y)
    import asyncio

    eg = ExamPaperGenerator("questions")
    # asyncio.run(eg.load_question_bank(1, 5))
    asyncio.run(eg.load_single_choice_bank())

    question_type = {'fill': 5, 'judge': 5, 'multiple_choice': 5, 'short_answer': 5, 'single_choice': 10}
    exam_paper_info = asyncio.run(eg.generate_exam(question_type))
    print(exam_paper_info)
