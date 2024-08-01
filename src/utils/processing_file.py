#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/03/15
# @Author  : MementoMori
# @File    : processing_file.py
# @Software: PyCharm

import os
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
from src.config.setting import settings
from src.utils.constant import COLUMNS_TO_KEEP


class GenerateFile:
    def __init__(self):
        pass

    @staticmethod
    def generate_xml(table_list, output_path=None, file_name=None):
        """
        生成xml
        Args:
            table_list:从数据库查询得到的一个二维列表,每行代表一条记录，每列对应一个字段
            output_path:
            file_name:

        Returns:

        """
        table_name = file_name.split(".")[0]
        root = ET.Element("root")
        for row in table_list:
            record = ET.SubElement(root, "record")
            for field_key, field_value in row.items():
                field_name = f"{field_key}"
                field_element = ET.SubElement(record, field_name)
                field_element.text = str(field_value)

        xml_data = ET.tostring(root, encoding="utf-8", method="xml").decode("utf-8")
        output_filename = os.path.join(output_path, table_name)
        # 写入到文件（可选）
        with open(output_filename, "w") as xml_file:
            xml_file.write(xml_data)

        return output_filename

    @staticmethod
    def generate_excel(table_list, zh_name=None, col_name=None, output_path=None, file_name=None):
        """
        生成excel
        Args:
            table_list: 从数据库查询得到的数据
            col_name: 列名列表
            output_path: 输出路径
            file_name: 文件名
            zh_name: 文件名
        Returns:
            输出文件的完整路径
        """
        table_name = file_name.split(".")[0]

        # 使用col_name作为列名创建DataFrame
        if col_name:
            if table_list and isinstance(table_list[0], dict):
                # 如果table_list是字典列表，则转换为DataFrame并使用col_name作为列名
                df = pd.DataFrame(table_list)
                df.columns = col_name
            else:
                # 如果table_list是其他类型，创建一个空的DataFrame并使用col_name作为列名
                df = pd.DataFrame(columns=col_name)
        elif table_name == settings.QUESTION_BANK:
            df = pd.DataFrame(table_list)[COLUMNS_TO_KEEP]
        else:
            df = pd.DataFrame(table_list)

        # 创建Workbook并添加Sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 添加列名到工作表
        ws.append(col_name)

        # 添加数据行到工作表
        for row in df.itertuples(index=False):
            ws.append(list(row))
        # 保存文件
        output_filename = os.path.join(output_path, zh_name)
        wb.save(output_filename)
        return output_filename

    @staticmethod
    def max_column_width(data, column_index):
        """
        获取某列中最长字符串的长度
        Args:
            data:
            column_index:

        Returns:

        """
        return max([len(str(row[column_index])) for row in data])

    def generate_pdf(self, table_list, zh_name=None, output_path=None, file_name=None):
        """
        生成pdf
        Args:
            table_list: 从数据库查询得到的数据
            output_path:
            file_name:
            zh_name:

        Returns:

        """
        pdf = FPDF("L")
        page_width = 270  # 单位毫米，此处仅为示例，实际值应为扣除页边距后的宽度

        simfang_path = "src/static/simfang.ttf"
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)  # 自动分页
        pdf.add_font('simfang', '', simfang_path, uni=True)
        pdf.set_font('simfang', '', 14)

        table_name = file_name.split(".")[0]
        column_titles = list(table_list[0].keys())
        user_data = [list(row.values()) for row in table_list]
        if table_name == settings.QUESTION_BANK:
            for idx, data in enumerate(user_data):
                question, answer = data[16], data[17]
                con = str(idx + 1) + ". " + question.strip()
                pdf.multi_cell(0, 10, txt=con)
        else:
            # 计算每列的最大宽度
            max_widths = [self.max_column_width(user_data, i) for i in range(len(column_titles))]
            column_widths = [int(page_width * w / sum(max_widths)) for w in max_widths]
            # 将数据写入PDF
            for i, (title, width) in enumerate(zip(column_titles, column_widths)):
                pdf.cell(width, 10, txt=title, border=1, align='C' if i == 0 else 'L', ln=(i == len(table_list[0]) - 1))
            for record in user_data:
                for col_idx, (value, width) in enumerate(zip(record, column_widths)):
                    pdf.cell(width, 10, txt=str(value), border=1, ln=(col_idx == len(record) - 1),
                             align='C' if col_idx == 0 else 'L')

        # 指定要保存的路径
        output_filename = os.path.join(output_path, zh_name)
        pdf.output(output_filename)
        return output_filename


class Plotter:
    def __init__(self, figure_size=(10, 6)):
        self.figure_size = figure_size

    @staticmethod
    def configure_plot(title, x_label=None, y_label=None):
        plt.title(title)
        if x_label:
            plt.xlabel(x_label)
        if y_label:
            plt.ylabel(y_label)
        plt.tight_layout()

    def create_line_plot(self, x_data, y_data, x_label='X-axis', y_label='Y-axis', title='Line Plot'):
        plt.figure(figsize=self.figure_size)
        plt.plot(x_data, y_data)
        self.configure_plot(title, x_label, y_label)
        plt.show()

    def create_scatter_plot(self, x_data, y_data, x_label='X-axis', y_label='Y-axis', title='Scatter Plot'):
        plt.figure(figsize=self.figure_size)
        plt.scatter(x_data, y_data)
        self.configure_plot(title, x_label, y_label)
        plt.show()

    def create_bar_chart(self, categories, values, x_label='Categories', y_label='Values', title='Bar Chart'):
        plt.figure(figsize=self.figure_size)
        plt.bar(categories, values)
        plt.xticks(rotation=45)
        self.configure_plot(title, x_label, y_label)
        plt.show()

    def create_pie_chart(self, sizes, labels, title='Pie Chart', autopct='%1.1f%%'):
        plt.figure(figsize=self.figure_size)
        plt.pie(sizes, labels=labels, autopct=autopct, startangle=90)
        plt.axis('equal')
        self.configure_plot(title)
        plt.show()

    def create_histogram(self, data, bins=10, x_label='Value Range', y_label='Frequency', title='Histogram'):
        plt.figure(figsize=self.figure_size)
        plt.hist(data, bins=bins, edgecolor='black')
        self.configure_plot(title, x_label, y_label)
        plt.show()


generate_file = GenerateFile()
# plotter = Plotter()

# if __name__ == '__main__':
# generate_xml()
# plotter = Plotter()
# # 示例数据
# x = range(10)
# y = [i ** 2 for i in x]
# categories = ['A', 'B', 'C', 'D', 'E']
# values = [30, 25, 40, 35, 20]
# sizes = [30, 25, 40, 35]
# labels = ['Category A', 'Category B', 'Category C', 'Category D']

# 使用示例
# plotter.create_line_plot(x, y)
# plotter.create_scatter_plot(x, y)
# plotter.create_bar_chart(categories, values)
# plotter.create_pie_chart(sizes, labels)
# plotter.create_histogram(y)
