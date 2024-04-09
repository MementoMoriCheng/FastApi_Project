import itertools
import os
import random
from collections import defaultdict
from datetime import datetime

# 四个列表
students = ['student1', 'student2', 'student3', 'student4', 'student5']
teachers = ['teacher1', 'teacher2', 'teacher3', 'teacher4']
planes = ['plane1', 'plane2', "plane3"]
routes = ['route1', 'route2', "route3", "route4"]


# 合并所有列表
# all_elements = students + teachers + planes + routes

# 进行全排列
# permutations_result = list(itertools.permutations(all_elements))
#
# # 输出结果
# for p in permutations_result:
#     print(p)

def gen_cartesian_product(students, teachers, planes):
    """
    资源全排列
    Args:
        students:
        teachers:
        planes:

    Returns:

    """

    return list(itertools.product(students, teachers, planes))


# dd_list = gen_cartesian_product( students, teachers, planes)
# print(dd_list)


def find_unique_elements(list1, list2):
    unique_elements = []

    for element2 in list2:
        is_unique = True

        for element1 in list1:
            if element2[0] == element1[0] or element2[1] == element1[1] or element2[2] == element1[2]:
                is_unique = False
                break

        if is_unique:
            unique_elements.append(element2)

    return unique_elements


list1 = [(6, 6, 4), (7, 7, 4), (5, 5, 5), (7, 7, 6)]
list2 = [(5, 5, 4), (5, 5, 5), (5, 5, 6), (5, 6, 4), (5, 6, 5), (5, 6, 6), (5, 7, 4), (5, 7, 5), (5, 7, 6), (5, 8, 4),
         (5, 8, 5), (5, 8, 6), (6, 5, 4), (6, 5, 5), (6, 5, 6), (6, 6, 4), (6, 6, 5), (6, 6, 6), (6, 7, 4), (6, 7, 5),
         (6, 7, 6), (6, 8, 4), (6, 8, 5), (6, 8, 6), (7, 5, 4), (7, 5, 5), (7, 5, 6), (7, 6, 4), (7, 6, 5), (7, 6, 6),
         (7, 7, 4), (7, 7, 5), (7, 7, 6), (7, 8, 4), (7, 8, 5), (7, 8, 6), (8, 5, 4), (8, 5, 5), (8, 5, 6), (8, 6, 4),
         (8, 6, 5), (8, 6, 6), (8, 7, 4), (8, 7, 5), (8, 7, 6), (8, 8, 4), (8, 8, 5), (8, 8, 6)]


# unique_elements = find_unique_elements(list1, list2)
# print(unique_elements)

# random.shuffle(list1)
# print(list1)


def generate_mysql_log_data(**log_data):
    """
    数据库数据格式
    Args:
        log_data:

    Returns:

    """
    log_records = {
        'id': 1,
        'level': 1,
        'is_delete': 0,
        'handle_user': log_data['handle_user'],
        'handle_reason': log_data['handle_reason'],
        'handle_params': 'Hello, UDP Server!',
        'entity_type': 'Hello, UDP Server!',
        'entity_id': 'Hello, UDP Server!',
        'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    return log_records


# a = generate_mysql_log_data(handle_user="李狗蛋", handle_reason="李狗蛋擦做")
# print(a)

d_ct = {'student_id': 5, 'route_id': 4, 'coach_id': 5, 'plane_id': 5, 'plan_duration': 120,
        'plan_time_start': '2024-03-06 13:00:00', 'plan_time_end': '2024-03-06 15:00:00',
        'real_time_start': '2024-03-06 13:00:00', 'real_time_end': '2024-03-06 15:00:00', 'handle_user': 'sysadmin',
        'handle_reason': None, 'status': 1, 'name': 'bbb', 'description': None,
        'expand_data': {'name': 'bbb', 'fly_route': {'table': 'fly_route', 'id': [4, 5, 6]}, 'status': 1,
                        'plane': {'table': 'plane', 'id': [4, 5, 6]}, 'coach': {'table': 'coach', 'id': [5, 6, 7, 8]},
                        'student': {'table': 'student', 'id': [5, 6, 7, 8]}, 'handle_user': 'sysadmin',
                        'description': None, 'plane_most_ply': 2, 'coach_most_ply': 5, 'student_most_ply': 10,
                        'flight_interval': 30, 'flight_duration': 120, 'start_time': '2024-03-06 08:00:00',
                        'end_time': '2024-03-06 18:00:00', 'exclude_time': []}}

class PlanObject:
    def __init__(self, data_dict):
        for key, value in data_dict.items():
            setattr(self, key, value)

# plan_obj = PlanObject(d_ct)
# print(plan_obj.student_id)
# list3 = [(6, 6, 4), (7, 7, 4), (5, 5, 5), (7, 7, 6)]
# for i, c in enumerate(list3):
#     list3.append(list1[i])
#     print(list3)
# local_path = "D:\Project\\flight-training"
# local_file_path = os.path.join(local_path, "gnss_data", "file_name")
#
# print(local_file_path)

from fpdf import FPDF

# # 假设这是您的数据
# data = [
#     ["sysadmin", 5, "秦英", "1981-10-04", "重庆 重庆市 渝北区", 3, 1, "430000202309276415", 1, "18978598195", 1],
#     ["sysadmin", 6, "江军", "1996-01-04", "天津 天津市 滨海新区", 4, 1, "440000197011271839", 1, "13247260713", 1],
#     ["sysadmin", 7, "吴洋", "1998-10-04", "福建省 南平市 顺昌县", 1, 1, "540000200405054632", 1, "13594052543", 1],
#     ["sysadmin", 8, "曾秀兰", "1997-01-15", "海南省 海口市 美兰区", 3, 1, "140000198906160383", 1, "18918593631", 2]
# ]
#
# # 定义列标题
# column_titles = ["职位", "员工ID", "姓名", "出生日期", "工作地点", "部门ID", "公司ID", "身份证号", "性别", "联系电话", "状态"]

def create_pdf(data, column_titles, output_filename):
    pdf = FPDF()

    # 添加一页
    simfang_path = "D:\\Project\\flight-training-server\\src\\static\\simfang.ttf"
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)  # 自动分页
    # pdf.set_font("Arial", size=12)
    pdf.add_font('simfang', '', simfang_path, uni=True)
    pdf.set_font('simfang', '', 12)

    # 设置字体
    # pdf.set_font("Arial", size=12)

    # 写入列标题
    for i, title in enumerate(column_titles):
        pdf.cell(50, 10, txt=title, border=1, align='C' if i == 0 else 'L', ln=(i == len(column_titles) - 1))

    # 写入每一行数据
    for row in data:
        for i, cell in enumerate(row):
            pdf.cell(50, 10, txt=str(cell), border=1, ln=(i == len(row) - 1), align='C' if i == 0 else 'L')

    # 保存到指定路径
    output_path = f"{output_filename}.pdf"
    pdf.output(output_path)

    print(f"PDF已成功保存至: {output_path}")

# 调用函数创建并保存PDF
# create_pdf(data, column_titles, "employee_data")

# from reportlab.lib import colors
# from reportlab.lib.pagesizes import letter
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
#
# # 创建PDF的基础结构
# pdf = SimpleDocTemplate("我的报告.pdf", pagesize=letter)
#
# # 创建表格数据
# data = [
#     ["课程", "教师", "时间", "地点"],
#     ["数学", "张老师", "8:00-10:00", "1号教室"],
#     ["语文", "李老师", "10:00-12:00", "2号教室"],
#     ["英语", "王老师", "14:00-16:00", "3号教室"],
# ]
#
# # 创建表格
# # t = Table(data)
#
# # 设置表格样式
# t.setStyle(TableStyle([
#     ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#     ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#     ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#     ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#     ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
# ]))
#
# # 将表格添加到PDF中
# # pdf.build([t])
from openpyxl import Workbook

wb = Workbook()
wb.save('file.xls')